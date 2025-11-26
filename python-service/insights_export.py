"""
Advanced Export System for Insights Agent
Generates PDF and Excel reports with embedded charts
"""

import pandas as pd
from typing import Dict, Any, List
from io import BytesIO
import base64

# PDF Generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️  reportlab not installed - PDF export disabled")

# Excel Generation
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed - Excel export disabled")


def create_pdf_report(result: Dict[str, Any], query: str, org_name: str = "Organization") -> bytes:
    """
    Generate a PDF report with the insights result.
    Returns PDF as bytes.
    """
    if not PDF_AVAILABLE:
        raise RuntimeError("PDF generation not available - install reportlab")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#14b8a6'),
        spaceAfter=30,
    )
    story.append(Paragraph(f"Insights Report - {org_name}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Query
    story.append(Paragraph(f"<b>Question:</b> {query}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Summary
    if result.get('summary'):
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#0ea5e9'),
            leftIndent=20,
            rightIndent=20,
            spaceAfter=20,
        )
        story.append(Paragraph(f"<i>{result['summary']}</i>", summary_style))
        story.append(Spacer(1, 0.3*inch))
    
    # Data visualization
    result_type = result.get('type', 'table')
    data = result.get('data', [])
    
    if result_type == 'table' and isinstance(data, list) and data:
        # Create table
        df = pd.DataFrame(data)
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#14b8a6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(t)
    
    elif result_type in ['bar_chart', 'line_chart'] and isinstance(data, list) and data:
        # Create chart
        drawing = Drawing(400, 200)
        
        if result_type == 'bar_chart':
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 50
            chart.height = 125
            chart.width = 300
            chart.data = [[item.get('y', 0) for item in data]]
            chart.categoryAxis.categoryNames = [str(item.get('x', ''))[:15] for item in data]
            chart.valueAxis.valueMin = 0
            chart.bars[0].fillColor = colors.HexColor('#14b8a6')
        else:  # line_chart
            chart = HorizontalLineChart()
            chart.x = 50
            chart.y = 50
            chart.height = 125
            chart.width = 300
            chart.data = [[item.get('y', 0) for item in data]]
            chart.categoryAxis.categoryNames = [str(item.get('x', ''))[:15] for item in data]
            chart.lines[0].strokeColor = colors.HexColor('#14b8a6')
            chart.lines[0].strokeWidth = 2
        
        drawing.add(chart)
        story.append(drawing)
    
    # Footer
    story.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,  # Center
    )
    from datetime import datetime
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by CoreMatch Insights Agent", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def create_excel_report(result: Dict[str, Any], query: str, org_name: str = "Organization") -> bytes:
    """
    Generate an Excel report with the insights result.
    Returns Excel file as bytes.
    """
    if not EXCEL_AVAILABLE:
        raise RuntimeError("Excel generation not available - install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Insights Report"
    
    # Header styling
    header_fill = PatternFill(start_color="14b8a6", end_color="14b8a6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    # Title
    ws['A1'] = f"Insights Report - {org_name}"
    ws['A1'].font = Font(bold=True, size=16, color="14b8a6")
    ws.merge_cells('A1:D1')
    
    # Query
    ws['A3'] = "Question:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = query
    
    # Summary
    if result.get('summary'):
        ws['A5'] = "Summary:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = result['summary']
        ws.merge_cells('B5:D5')
        ws['B5'].alignment = Alignment(wrap_text=True)
    
    # Data
    result_type = result.get('type', 'table')
    data = result.get('data', [])
    
    start_row = 7
    
    if isinstance(data, list) and data:
        df = pd.DataFrame(data)
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add chart if applicable
        if result_type in ['bar_chart', 'line_chart', 'pie_chart'] and len(df) > 0:
            chart_row = start_row + len(df) + 3
            
            if result_type == 'bar_chart':
                chart = BarChart()
                chart.title = "Data Visualization"
                chart.style = 10
                chart.y_axis.title = df.columns[1] if len(df.columns) > 1 else 'Value'
                chart.x_axis.title = df.columns[0]
                
                data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(df))
                cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(df))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                ws.add_chart(chart, f"A{chart_row}")
            
            elif result_type == 'line_chart':
                chart = LineChart()
                chart.title = "Trend Analysis"
                chart.style = 10
                chart.y_axis.title = df.columns[1] if len(df.columns) > 1 else 'Value'
                chart.x_axis.title = df.columns[0]
                
                data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(df))
                cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(df))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                ws.add_chart(chart, f"A{chart_row}")
            
            elif result_type == 'pie_chart':
                chart = PieChart()
                chart.title = "Distribution"
                
                data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(df))
                cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(df))
                chart.add_data(data_ref)
                chart.set_categories(cats_ref)
                
                ws.add_chart(chart, f"A{chart_row}")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_to_base64(file_bytes: bytes, file_type: str) -> str:
    """
    Convert file bytes to base64 for frontend download.
    """
    b64 = base64.b64encode(file_bytes).decode('utf-8')
    mime_type = 'application/pdf' if file_type == 'pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return f"data:{mime_type};base64,{b64}"
